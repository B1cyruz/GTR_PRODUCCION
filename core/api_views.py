import json
from datetime import datetime
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden, HttpResponseNotFound
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Q

from core.models import (
    User, Driver, Client, DeliveryPoint, DeliveryPointHistory,
    Delivery, Route, RouteStop, AuditLog, SystemSetting
)
from core.services.geo import (
    haversine_distance_meters,
    find_nearby_delivery_points,
    save_base64_image
)
from core.services.distribution import distribute_random_deliveries
from core.services.audit import log_audit_event, get_client_ip
from django.contrib.auth import authenticate, login, logout

# --- API DE AUTENTICACIÓN AJAX & SSO ---
@csrf_exempt
def api_auth_login(request):
    if request.method != 'POST':
        return HttpResponseBadRequest()
    try:
        body = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest(json.dumps({"detail": "JSON inválido"}), content_type="application/json")

    email_or_user = (body.get('email_or_user') or body.get('email') or body.get('username') or '').strip()
    password = body.get('password', '').strip()

    user_obj = User.objects.filter(Q(username__iexact=email_or_user) | Q(email__iexact=email_or_user)).first()
    if not user_obj:
        return JsonResponse({"detail": "Usuario no encontrado", "success": False}, status=400)

    user = authenticate(request, username=user_obj.username, password=password)
    if not user:
        return JsonResponse({"detail": "Contraseña incorrecta", "success": False}, status=400)

    if not user.is_active:
        return JsonResponse({"detail": "Usuario desactivado", "success": False}, status=403)

    login(request, user)
    user.last_login = timezone.now()
    user.save(update_fields=['last_login'])

    log_audit_event(
        user=user,
        action="API_LOGIN",
        module="AUTH",
        ip_address=get_client_ip(request),
        details={"role": user.role}
    )

    redirect_url = '/repartidor/ruta-activa/' if user.role == 'REPARTIDOR' else '/dashboard/'
    return JsonResponse({
        "success": True,
        "message": f"Bienvenido, {user.full_name}",
        "role": user.role,
        "redirect_url": redirect_url
    })


@csrf_exempt
def api_auth_provider_login(request):
    if request.method != 'POST':
        return HttpResponseBadRequest()
    try:
        body = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest()

    provider = body.get('provider', 'demo_root').strip()

    target_user = None
    if provider in ['admin', 'root', 'demo_root']:
        target_user = User.objects.filter(username__in=['admin', 'root']).first() or User.objects.filter(role='ROOT').first()
    elif provider in ['coordinador', 'demo_coord']:
        target_user = User.objects.filter(username='coordinador').first() or User.objects.filter(role='COORDINADOR').first()
    elif provider in ['repartidor1', 'repartidor2', 'repartidor3', 'repartidor4', 'demo_driver']:
        if provider.startswith('repartidor'):
            target_user = User.objects.filter(username=provider).first()
        if not target_user:
            target_user = User.objects.filter(role='REPARTIDOR').first()
    else:
        email = body.get('email')
        if email:
            target_user = User.objects.filter(Q(email__iexact=email) | Q(username__iexact=email)).first()
        if not target_user:
            target_user = User.objects.filter(role='ROOT').first()

    if not target_user:
        return JsonResponse({"detail": "Perfil demo no disponible. Ejecuta python manage.py seed_data.", "success": False}, status=404)

    # Login directo de demo
    login(request, target_user)
    target_user.last_login = timezone.now()
    target_user.save(update_fields=['last_login'])

    log_audit_event(
        user=target_user,
        action="DEMO_SWITCH_LOGIN",
        module="AUTH",
        ip_address=get_client_ip(request),
        details={"provider": provider, "role": target_user.role}
    )

    redirect_url = '/repartidor/ruta-activa/' if target_user.role == 'REPARTIDOR' else '/dashboard/'
    return JsonResponse({
        "success": True,
        "message": f"Sesión iniciada como {target_user.full_name} ({target_user.role})",
        "role": target_user.role,
        "redirect_url": redirect_url
    })


# --- HELPER DE SERIALIZACIÓN DE PUNTO GPS ---
def serialize_delivery_point(p: DeliveryPoint) -> dict:
    return {
        "id": p.id,
        "name": p.name,
        "address": p.address,
        "neighborhood": p.neighborhood or "Cartagena",
        "city": p.city or "Cartagena",
        "reference_point": p.reference_point,
        "latitude": p.latitude,
        "longitude": p.longitude,
        "gps_accuracy": p.gps_accuracy,
        "location_source": p.location_source,
        "status": p.status,
        "notes": p.notes,
        "photo_url": p.photo_url,
        "client_id": p.client_id,
        "client": {
            "id": p.client.id,
            "name": p.client.name,
            "phone": p.client.phone,
            "document_id": p.client.document_id,
            "city": p.client.city
        } if p.client else None,
        "created_by_user_id": p.created_by_user_id,
        "created_by_user_name": p.created_by_user_name,
        "created_by_user_role": p.created_by_user_role,
        "created_at": p.created_at.isoformat() if p.created_at else None,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None
    }


# --- API DE PUNTOS GPS (BASE GENERAL) ---
@csrf_exempt
def api_points_list_create(request):
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No autenticado"}, status=401)

    if request.method == 'GET':
        search = request.GET.get('search', '').strip()
        neighborhood = request.GET.get('neighborhood', '').strip()
        client_id = request.GET.get('client_id', '').strip()
        status_filter = request.GET.get('status_filter', 'ACTIVO').strip()
        limit = int(request.GET.get('limit', 100))

        qs = DeliveryPoint.objects.select_related('client', 'created_by_user').all()
        if status_filter and status_filter.upper() != 'ALL':
            qs = qs.filter(status=status_filter.upper())
        if client_id and client_id.isdigit():
            qs = qs.filter(client_id=int(client_id))
        if neighborhood:
            qs = qs.filter(neighborhood__icontains=neighborhood)
        if search:
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(address__icontains=search) |
                Q(reference_point__icontains=search) |
                Q(client__name__icontains=search) |
                Q(client__document_id__icontains=search)
            )

        points = qs.order_by('-created_at')[:limit]
        return JsonResponse([serialize_delivery_point(p) for p in points], safe=False)

    elif request.method == 'POST':
        try:
            body = json.loads(request.body.decode('utf-8'))
        except Exception:
            return HttpResponseBadRequest(json.dumps({"detail": "JSON inválido"}), content_type="application/json")

        name = body.get('name', '').strip()
        address = body.get('address', '').strip()
        if not name or not address:
            return JsonResponse({"detail": "El nombre y la dirección son obligatorios"}, status=400)

        latitude = float(body.get('latitude', 0.0))
        longitude = float(body.get('longitude', 0.0))
        gps_accuracy = body.get('gps_accuracy')
        if gps_accuracy is not None:
            gps_accuracy = float(gps_accuracy)
        
        force_duplicate = body.get('force_create_duplicate', False)
        target_client_id = body.get('client_id')
        client_name = body.get('client_name')
        client_phone = body.get('client_phone')
        client_doc = body.get('client_document')

        # Buscar o crear cliente
        if not target_client_id and (client_name or client_doc or client_phone):
            client = None
            if client_doc:
                client = Client.objects.filter(document_id=client_doc.strip()).first()
            if not client and client_phone:
                client = Client.objects.filter(phone=client_phone.strip()).first()
            if not client and client_name:
                client = Client.objects.filter(name__iexact=client_name.strip()).first()

            if not client:
                client = Client.objects.create(
                    name=client_name.strip() if client_name else "Cliente General",
                    document_id=client_doc.strip() if client_doc else None,
                    phone=client_phone.strip() if client_phone else "3000000000",
                    address=address,
                    neighborhood=body.get('neighborhood', 'Cartagena'),
                    city=body.get('city', 'Cartagena'),
                    created_by_user=request.user
                )
            target_client_id = client.id

        # Validación de duplicados por proximidad GPS (<=50m) o misma dirección
        if not force_duplicate:
            nearby = find_nearby_delivery_points(latitude, longitude, radius_meters=50.0)
            
            # Comprobar duplicados por dirección
            match_addr = DeliveryPoint.objects.filter(address__iexact=address).exclude(status='INACTIVO')
            for m in match_addr:
                if not any(n.get("point_id") == m.id for n in nearby):
                    dist = haversine_distance_meters(latitude, longitude, m.latitude, m.longitude)
                    nearby.append({
                        "point_id": m.id,
                        "name": m.name,
                        "address": m.address,
                        "neighborhood": m.neighborhood,
                        "city": m.city,
                        "reference_point": m.reference_point,
                        "latitude": m.latitude,
                        "longitude": m.longitude,
                        "distance_meters": round(dist, 1),
                        "photo_url": m.photo_url,
                        "client_id": m.client_id,
                        "client_name": m.client.name if m.client else None,
                        "client_document": m.client.document_id if m.client else None,
                        "client_phone": m.client.phone if m.client else None,
                        "updated_at": (m.updated_at or m.created_at).isoformat() if (m.updated_at or m.created_at) else None
                    })

            if nearby:
                return JsonResponse({
                    "message": "Se encontró un punto registrado cerca de esta ubicación (menos de 50m) o con la misma dirección.",
                    "duplicate_warning": True,
                    "nearby_points": nearby
                }, status=409)

        # Procesar fotografía capturada con la cámara
        photo_data = body.get('photo_data')
        saved_photo_url = save_base64_image(photo_data, subfolder="points", prefix="point_") or body.get('photo_url')

        new_point = DeliveryPoint.objects.create(
            client_id=target_client_id,
            name=name,
            address=address,
            neighborhood=body.get('neighborhood', 'Cartagena'),
            city=body.get('city', 'Cartagena'),
            reference_point=body.get('reference_point'),
            latitude=latitude,
            longitude=longitude,
            gps_accuracy=gps_accuracy,
            location_source=body.get('location_source', 'GPS_DEVICE'),
            status=body.get('status', 'ACTIVO'),
            notes=body.get('notes'),
            photo_url=saved_photo_url,
            created_by_user=request.user,
            updated_by_user=request.user
        )

        delivery_id = body.get('delivery_id')
        if delivery_id:
            deliv = Delivery.objects.filter(id=delivery_id).first()
            if deliv:
                deliv.delivery_point = new_point
                deliv.save(update_fields=['delivery_point'])

        # Registro en el historial
        DeliveryPointHistory.objects.create(
            delivery_point=new_point,
            user=request.user,
            action="CREACION",
            new_latitude=latitude,
            new_longitude=longitude,
            gps_accuracy=gps_accuracy,
            photo_url=saved_photo_url,
            notes=f"Punto registrado por {request.user.full_name} ({request.user.role}). Fuente: {new_point.location_source}"
        )

        # Auditoría
        log_audit_event(
            user=request.user,
            action="CREATE_GPS_POINT",
            module="POINTS",
            target_id=str(new_point.id),
            ip_address=get_client_ip(request),
            details={
                "name": new_point.name,
                "lat": new_point.latitude,
                "lng": new_point.longitude,
                "photo_url": new_point.photo_url,
                "driver_name": request.user.full_name
            }
        )

        return JsonResponse(serialize_delivery_point(new_point), status=201)

    return HttpResponseBadRequest()


@csrf_exempt
def api_points_nearby(request):
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No autenticado"}, status=401)

    try:
        lat = float(request.GET.get('latitude'))
        lng = float(request.GET.get('longitude'))
        radius = float(request.GET.get('radius', 50.0))
    except (TypeError, ValueError):
        return JsonResponse({"detail": "Coordenadas latitud y longitud requeridas"}, status=400)

    nearby = find_nearby_delivery_points(lat, lng, radius_meters=radius)
    return JsonResponse(nearby, safe=False)


@csrf_exempt
def api_point_detail_update(request, point_id):
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No autenticado"}, status=401)

    point = DeliveryPoint.objects.filter(id=point_id).first()
    if not point:
        return JsonResponse({"detail": "Punto de entrega no encontrado"}, status=404)

    if request.method == 'GET':
        return JsonResponse(serialize_delivery_point(point))

    elif request.method in ['PUT', 'PATCH']:
        try:
            body = json.loads(request.body.decode('utf-8'))
        except Exception:
            return HttpResponseBadRequest(json.dumps({"detail": "JSON inválido"}), content_type="application/json")

        old_lat = point.latitude
        old_lng = point.longitude
        gps_changed = False

        if 'name' in body:
            point.name = body['name'].strip()
        if 'address' in body:
            point.address = body['address'].strip()
        if 'neighborhood' in body:
            point.neighborhood = body['neighborhood'].strip()
        if 'city' in body:
            point.city = body['city'].strip()
        if 'reference_point' in body:
            point.reference_point = body['reference_point']
        if 'client_id' in body:
            point.client_id = body['client_id']
        if 'status' in body:
            point.status = body['status']
        if 'notes' in body:
            point.notes = body['notes']

        # Si se envía una nueva fotografía desde la cámara
        if 'photo_data' in body and body['photo_data']:
            saved_photo = save_base64_image(body['photo_data'], subfolder="points", prefix=f"point_{point.id}_")
            if saved_photo:
                point.photo_url = saved_photo
        elif 'photo_url' in body:
            point.photo_url = body['photo_url']

        if 'latitude' in body and 'longitude' in body:
            new_lat = float(body['latitude'])
            new_lng = float(body['longitude'])
            if old_lat != new_lat or old_lng != new_lng:
                gps_changed = True
                point.latitude = new_lat
                point.longitude = new_lng
                if 'gps_accuracy' in body and body['gps_accuracy'] is not None:
                    point.gps_accuracy = float(body['gps_accuracy'])
                if 'location_source' in body:
                    point.location_source = body['location_source']

        point.updated_by_user = request.user
        point.save()

        # Registrar en Historial
        DeliveryPointHistory.objects.create(
            delivery_point=point,
            user=request.user,
            action="ACTUALIZACION_GPS" if gps_changed else "EDICION_DATOS",
            previous_latitude=old_lat,
            previous_longitude=old_lng,
            new_latitude=point.latitude,
            new_longitude=point.longitude,
            gps_accuracy=point.gps_accuracy,
            photo_url=point.photo_url,
            notes=f"Actualizado por {request.user.full_name} ({request.user.role}). Notas: {body.get('notes', 'Actualización de datos y coordenadas')}"
        )

        delivery_id = body.get('delivery_id')
        if delivery_id:
            deliv = Delivery.objects.filter(id=delivery_id).first()
            if deliv:
                deliv.delivery_point = point
                deliv.save(update_fields=['delivery_point'])

        log_audit_event(
            user=request.user,
            action="UPDATE_GPS_POINT",
            module="POINTS",
            target_id=str(point.id),
            ip_address=get_client_ip(request),
            details={
                "name": point.name,
                "lat": point.latitude,
                "lng": point.longitude,
                "photo_url": point.photo_url,
                "gps_changed": gps_changed,
                "driver_name": request.user.full_name
            }
        )

        return JsonResponse(serialize_delivery_point(point))

    return HttpResponseBadRequest()


@csrf_exempt
def api_point_history(request, point_id):
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No autenticado"}, status=401)

    point = DeliveryPoint.objects.filter(id=point_id).first()
    if not point:
        return JsonResponse({"detail": "Punto de entrega no encontrado"}, status=404)

    history = DeliveryPointHistory.objects.filter(delivery_point=point).select_related('user').order_by('-created_at')
    results = []
    for h in history:
        results.append({
            "id": h.id,
            "delivery_point_id": h.delivery_point_id,
            "user_id": h.user_id,
            "user_name": h.user_name,
            "action": h.action,
            "previous_latitude": h.previous_latitude,
            "previous_longitude": h.previous_longitude,
            "new_latitude": h.new_latitude,
            "new_longitude": h.new_longitude,
            "gps_accuracy": h.gps_accuracy,
            "photo_url": h.photo_url,
            "notes": h.notes,
            "created_at": h.created_at.isoformat()
        })
    return JsonResponse(results, safe=False)


# --- API DE RUTAS Y PARADAS ---
@csrf_exempt
def api_update_stop_status(request, stop_id):
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No autenticado"}, status=401)

    stop = RouteStop.objects.filter(id=stop_id).select_related('route', 'delivery').first()
    if not stop:
        return JsonResponse({"detail": "Parada no encontrada"}, status=404)

    if request.user.role == 'REPARTIDOR' and request.user.driver_id and stop.route.driver_id != request.user.driver_id:
        return JsonResponse({"detail": "No puedes modificar paradas de otro conductor"}, status=403)

    body = {}
    if request.body:
        try:
            body = json.loads(request.body.decode('utf-8'))
        except Exception:
            pass

    target_status = body.get('status_str') or request.GET.get('status_str') or 'COMPLETADO'
    target_notes = body.get('notes') or request.GET.get('notes')
    target_id_card = body.get('recipient_id_card')
    target_recipient_name = body.get('recipient_name')
    target_photo = body.get('proof_photo_data')

    stop.status = target_status
    if target_notes:
        stop.notes = target_notes

    if target_status == 'COMPLETADO':
        stop.completion_time = timezone.now()
        stop.delivery.status = 'ENTREGADO'
        stop.delivery.delivered_at = timezone.now()
        if target_id_card:
            stop.delivery.recipient_id_card = target_id_card.strip()
        if target_recipient_name:
            stop.delivery.recipient_name = target_recipient_name.strip()
        
        target_ref_point = body.get('reference_point')
        if target_ref_point is not None:
            stop.delivery.reference_point = target_ref_point.strip()

        if target_photo:
            saved_proof = save_base64_image(target_photo, subfolder="proofs", prefix=f"proof_{stop.delivery_id}_")
            if saved_proof:
                stop.delivery.proof_photo_url = saved_proof

        # Manejo de coordenadas GPS al momento de entrega
        delivery_lat = body.get('delivery_lat')
        delivery_lng = body.get('delivery_lng')
        gps_accuracy = body.get('gps_accuracy')
        if delivery_lat is not None and delivery_lng is not None:
            try:
                lat_f = float(delivery_lat)
                lng_f = float(delivery_lng)
                if lat_f != 0.0 and lng_f != 0.0:
                    stop.delivery.latitude = lat_f
                    stop.delivery.longitude = lng_f
                    if stop.route and stop.route.driver:
                        stop.route.driver.current_lat = lat_f
                        stop.route.driver.current_lng = lng_f
                        stop.route.driver.save(update_fields=['current_lat', 'current_lng'])
            except (ValueError, TypeError):
                pass

        # Registro o vinculación en la Base General de Puntos GPS si no estaba previamente registrado
        register_gps = body.get('register_new_point', False)
        point_name = body.get('point_name') or stop.delivery.place_name or stop.delivery.recipient_name
        
        if register_gps and not stop.delivery.delivery_point:
            lat_check = stop.delivery.latitude
            lng_check = stop.delivery.longitude
            existing_nearby = find_nearby_delivery_points(lat_check, lng_check, radius_meters=50.0)
            
            if existing_nearby:
                # Si ya existe previamente, vincular y omitir duplicación
                matched_id = existing_nearby[0]["point_id"]
                matched_point = DeliveryPoint.objects.filter(id=matched_id).first()
                if matched_point:
                    stop.delivery.delivery_point = matched_point
                    if not matched_point.photo_url and stop.delivery.proof_photo_url:
                        matched_point.photo_url = stop.delivery.proof_photo_url
                        matched_point.save(update_fields=['photo_url'])
            else:
                # Crear nuevo punto GPS en base general
                client_obj = stop.delivery.client
                if not client_obj:
                    client_obj = Client.objects.create(
                        name=point_name,
                        phone=stop.delivery.recipient_phone,
                        email=stop.delivery.recipient_email,
                        address=stop.delivery.address,
                        neighborhood=stop.delivery.neighborhood,
                        city=stop.delivery.city,
                        created_by_user=request.user
                    )
                    stop.delivery.client = client_obj

                new_point = DeliveryPoint.objects.create(
                    client=client_obj,
                    name=point_name,
                    address=stop.delivery.address,
                    neighborhood=stop.delivery.neighborhood,
                    city=stop.delivery.city,
                    reference_point=stop.delivery.reference_point,
                    latitude=lat_check,
                    longitude=lng_check,
                    gps_accuracy=float(gps_accuracy) if gps_accuracy else None,
                    location_source='GPS_MOBILE_DRIVER',
                    photo_url=stop.delivery.proof_photo_url,
                    created_by_user=request.user,
                    notes=f"Registrado desde entrega móvil por {request.user.full_name}"
                )
                stop.delivery.delivery_point = new_point

                DeliveryPointHistory.objects.create(
                    delivery_point=new_point,
                    user=request.user,
                    action='CREACION_EN_ENTREGA',
                    new_latitude=lat_check,
                    new_longitude=lng_check,
                    gps_accuracy=float(gps_accuracy) if gps_accuracy else None,
                    photo_url=stop.delivery.proof_photo_url,
                    notes=f"Punto creado automáticamente durante la entrega del envío {stop.delivery.tracking_code}"
                )

        stop.delivery.save()

    elif target_status == 'EN_CAMINO':
        stop.delivery.status = 'EN_CAMINO'
        stop.delivery.save(update_fields=['status'])
    elif target_status == 'INCIDENCIA':
        stop.delivery.status = 'INCIDENCIA'
        if target_notes:
            stop.delivery.incident_reason = target_notes
        stop.delivery.save()

    stop.save()

    # Recalcular conteo de paradas completadas
    completed_count = RouteStop.objects.filter(route=stop.route, status='COMPLETADO').count()
    stop.route.completed_stops = completed_count
    if completed_count == stop.route.total_stops and stop.route.total_stops > 0:
        stop.route.status = 'FINALIZADA'
        stop.route.finished_at = timezone.now()
    stop.route.save()

    return JsonResponse({
        "success": True,
        "stop_id": stop.id,
        "status": stop.status,
        "completed_stops": stop.route.completed_stops,
        "total_stops": stop.route.total_stops
    })


@csrf_exempt
def api_distribute_random_routes(request):
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No autenticado"}, status=401)

    if request.user.role == 'REPARTIDOR':
        return JsonResponse({"detail": "Acceso denegado: solo Coordinador y Root pueden distribuir rutas"}, status=403)

    try:
        result = distribute_random_deliveries(seed_if_needed=True)
        log_audit_event(
            user=request.user,
            action="DISTRIBUTE_RANDOM_ROUTES",
            module="ROUTES",
            ip_address=get_client_ip(request),
            details={"drivers_allocated": result["drivers_allocated"]}
        )
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({"detail": str(e)}, status=500)


# --- API DE USUARIOS (ROOT ONLY) ---
@csrf_exempt
def api_users_list_create(request):
    if not request.user.is_authenticated or request.user.role != 'ROOT':
        return JsonResponse({"detail": "Acceso exclusivo para ROOT"}, status=403)

    if request.method == 'GET':
        users = User.objects.select_related('driver').order_by('id')
        data = []
        for u in users:
            data.append({
                "id": u.id,
                "username": u.username,
                "email": u.email,
                "full_name": u.full_name,
                "role": u.role,
                "provider": u.provider,
                "phone": u.phone,
                "is_active": u.is_active,
                "driver_id": u.driver_id,
                "driver_name": u.driver.name if u.driver else None,
                "last_login": u.last_login.isoformat() if u.last_login else None
            })
        return JsonResponse(data, safe=False)

    elif request.method == 'POST':
        try:
            body = json.loads(request.body.decode('utf-8'))
        except Exception:
            return HttpResponseBadRequest()

        email = body.get('email', '').strip()
        full_name = body.get('full_name', '').strip()
        role = body.get('role', 'COORDINADOR')
        password = body.get('password', 'GTRPassword2026!')

        if User.objects.filter(Q(username=email) | Q(email=email)).exists():
            return JsonResponse({"detail": "Ya existe un usuario con este correo electrónico"}, status=400)

        user = User.objects.create_user(
            username=email,
            email=email,
            password=password,
            full_name=full_name,
            role=role,
            phone=body.get('phone'),
            driver_id=body.get('driver_id')
        )
        return JsonResponse({"id": user.id, "email": user.email, "full_name": user.full_name, "role": user.role}, status=201)

    return HttpResponseBadRequest()


# --- API DE AUDITORÍA (ROOT ONLY) ---
@csrf_exempt
def api_audit_list(request):
    if not request.user.is_authenticated or request.user.role != 'ROOT':
        return JsonResponse({"detail": "Acceso exclusivo para ROOT"}, status=403)

    logs = AuditLog.objects.select_related('user').order_by('-created_at')[:200]
    data = []
    for l in logs:
        data.append({
            "id": l.id,
            "user_id": l.user_id,
            "user_name": l.user_name,
            "action": l.action,
            "module": l.module,
            "target_id": l.target_id,
            "ip_address": l.ip_address,
            "details": l.details_dict,
            "status": l.status,
            "created_at": l.created_at.isoformat()
        })
    return JsonResponse(data, safe=False)


# ==============================================================================
# --- MÓDULO: CARGA DE DATOS MULTIFORMATO & ESCÁNER QR / CÓDIGOS DE BARRA ---
# ==============================================================================

CARTAGENA_DEFAULT_COORDS = {
    'centro': (10.4236, -75.5524),
    'centro histórico': (10.4236, -75.5524),
    'getsemaní': (10.4208, -75.5458),
    'getsemani': (10.4208, -75.5458),
    'san diego': (10.4260, -75.5475),
    'bocagrande': (10.4012, -75.5543),
    'castillogrande': (10.3925, -75.5505),
    'el laguito': (10.3900, -75.5580),
    'manga': (10.4132, -75.5367),
    'pie de la popa': (10.4185, -75.5298),
    'la popa': (10.4190, -75.5260),
    'torices': (10.4320, -75.5340),
    'crespo': (10.4485, -75.5180),
    'el cabrero': (10.4310, -75.5420),
    'marbella': (10.4360, -75.5380),
    'los alpes': (10.3950, -75.4920),
    'el bosque': (10.3880, -75.5150),
    'mamonal': (10.3350, -75.5050),
    'chambacú': (10.4245, -75.5410),
    'cartagena': (10.4100, -75.5350)
}

def resolve_cartagena_coords(neighborhood, address=None):
    if neighborhood:
        n_clean = str(neighborhood).strip().lower()
        for key, coords in CARTAGENA_DEFAULT_COORDS.items():
            if key in n_clean or n_clean in key:
                return coords
    if address:
        a_clean = str(address).strip().lower()
        for key, coords in CARTAGENA_DEFAULT_COORDS.items():
            if key in a_clean:
                return coords
    return (10.4100, -75.5350)


@csrf_exempt
def api_bulk_import_deliveries(request):
    """
    Importa masivamente encomiendas y paquetes desde JSON o archivo Excel (.xlsx/.xls), CSV.
    """
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No autenticado", "success": False}, status=401)

    import io
    import csv
    records_to_process = []

    if request.method == 'POST':
        # Caso 1: Archivo subido por multipart/form-data
        if request.FILES.get('file'):
            uploaded_file = request.FILES['file']
            file_name = uploaded_file.name.lower()
            
            try:
                if file_name.endswith('.json'):
                    content = uploaded_file.read().decode('utf-8')
                    records_to_process = json.loads(content)
                    if isinstance(records_to_process, dict) and 'deliveries' in records_to_process:
                        records_to_process = records_to_process['deliveries']

                elif file_name.endswith('.csv') or file_name.endswith('.txt'):
                    content = uploaded_file.read().decode('utf-8-sig', errors='ignore')
                    reader = csv.DictReader(io.StringIO(content))
                    records_to_process = list(reader)

                elif file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                    import openpyxl
                    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                    ws = wb.active
                    headers = [str(cell.value or '').strip() for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(row):
                            row_dict = {}
                            for h, val in zip(headers, row):
                                if h:
                                    row_dict[h] = val
                            records_to_process.append(row_dict)

                else:
                    return JsonResponse({"detail": "Formato de archivo no soportado. Use .xlsx, .csv o .json", "success": False}, status=400)

            except Exception as e:
                return JsonResponse({"detail": f"Error al procesar archivo: {str(e)}", "success": False}, status=400)

        # Caso 2: Payload JSON con lista de registros
        else:
            try:
                body = json.loads(request.body.decode('utf-8'))
                if isinstance(body, list):
                    records_to_process = body
                elif isinstance(body, dict):
                    records_to_process = body.get('deliveries') or body.get('items') or [body]
            except Exception as e:
                return JsonResponse({"detail": f"Payload JSON inválido: {str(e)}", "success": False}, status=400)

        if not records_to_process:
            return JsonResponse({"detail": "No se encontraron registros para importar", "success": False}, status=400)

        created_count = 0
        updated_count = 0
        errors = []
        created_ids = []

        import random

        for idx, row in enumerate(records_to_process, start=1):
            try:
                # Mapeo flexible de columnas
                def get_field(keys, default=''):
                    for k in keys:
                        for row_key in row.keys():
                            if str(row_key).strip().lower() == str(k).strip().lower():
                                val = row.get(row_key)
                                if val is not None and str(val).strip() != '':
                                    return val
                    return default

                tracking = str(get_field(['tracking_code', 'guia', 'tracking', 'id_guia', 'codigo'], '')).strip()
                if not tracking:
                    tracking = f"GTR-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"

                recipient_name = str(get_field(['recipient_name', 'destinatario', 'cliente', 'nombre'], 'Destinatario')).strip()
                recipient_phone = str(get_field(['recipient_phone', 'telefono', 'celular', 'phone', 'contacto'], '3000000000')).strip()
                recipient_email = str(get_field(['recipient_email', 'email', 'correo'], '')).strip() or None

                sender_name = str(get_field(['sender_name', 'remitente', 'empresa_origen', 'sender'], '')).strip() or None
                sender_phone = str(get_field(['sender_phone', 'telefono_remitente', 'tel_remitente'], '')).strip() or None
                sender_doc = str(get_field(['sender_document', 'documento_remitente', 'nit_remitente'], '')).strip() or None

                address = str(get_field(['address', 'direccion', 'dir_entrega', 'ubicacion'], 'Cartagena')).strip()
                neighborhood = str(get_field(['neighborhood', 'barrio', 'zona', 'sector'], 'Centro Histórico')).strip()
                city = str(get_field(['city', 'ciudad'], 'Cartagena')).strip()
                reference_point = str(get_field(['reference_point', 'punto_referencia', 'referencia'], '')).strip() or None

                pkg_type = str(get_field(['package_type', 'tipo_envio', 'tipo_paquete', 'tipo'], 'PAQUETE_MEDIANO')).strip().upper().replace(' ', '_')
                
                try:
                    weight = float(get_field(['weight_kg', 'peso_kg', 'peso', 'weight'], 1.0))
                except (ValueError, TypeError):
                    weight = 1.0

                priority = str(get_field(['priority', 'prioridad'], 'MEDIA')).strip().upper()
                if priority not in ['ALTA', 'MEDIA', 'BAJA']:
                    priority = 'MEDIA'

                notes = str(get_field(['notes', 'observaciones', 'notas', 'indicaciones'], '')).strip() or None

                # Coordenadas
                lat_val = get_field(['latitude', 'lat', 'latitud'], None)
                lng_val = get_field(['longitude', 'lng', 'longitud'], None)
                
                if lat_val is not None and lng_val is not None:
                    try:
                        lat_f = float(lat_val)
                        lng_f = float(lng_val)
                    except (ValueError, TypeError):
                        lat_f, lng_f = resolve_cartagena_coords(neighborhood, address)
                else:
                    lat_f, lng_f = resolve_cartagena_coords(neighborhood, address)

                # Crear o vincular cliente
                client_obj, _ = Client.objects.get_or_create(
                    name=recipient_name,
                    phone=recipient_phone,
                    defaults={
                        "email": recipient_email,
                        "address": address,
                        "neighborhood": neighborhood,
                        "city": city,
                        "created_by_user": request.user
                    }
                )

                # Crear o actualizar Delivery
                delivery, created = Delivery.objects.update_or_create(
                    tracking_code=tracking,
                    defaults={
                        "client": client_obj,
                        "recipient_name": recipient_name,
                        "recipient_phone": recipient_phone,
                        "recipient_email": recipient_email,
                        "sender_name": sender_name,
                        "sender_phone": sender_phone,
                        "sender_document": sender_doc,
                        "address": address,
                        "neighborhood": neighborhood,
                        "city": city,
                        "latitude": lat_f,
                        "longitude": lng_f,
                        "reference_point": reference_point,
                        "package_type": pkg_type,
                        "weight_kg": weight,
                        "priority": priority,
                        "notes": notes,
                        "status": "PENDIENTE"
                    }
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

                created_ids.append({
                    "id": delivery.id,
                    "tracking_code": delivery.tracking_code,
                    "recipient_name": delivery.recipient_name,
                    "address": delivery.address,
                    "neighborhood": delivery.neighborhood,
                    "priority": delivery.priority
                })

            except Exception as row_err:
                errors.append(f"Fila {idx}: {str(row_err)}")

        log_audit_event(
            user=request.user,
            action="BULK_IMPORT_DELIVERIES",
            module="DISPATCH",
            ip_address=get_client_ip(request),
            details={"created": created_count, "updated": updated_count, "errors_count": len(errors)}
        )

        return JsonResponse({
            "success": True,
            "created_count": created_count,
            "updated_count": updated_count,
            "total_processed": len(records_to_process),
            "errors": errors,
            "items": created_ids[:50]
        })

    return HttpResponseBadRequest()


@csrf_exempt
def api_download_template(request):
    """
    Descarga la plantilla oficial de importación en formato Excel (.xlsx) o CSV.
    """
    from django.http import HttpResponse
    import io

    file_format = request.GET.get('format', 'xlsx').lower()

    sample_headers = [
        "tracking_code", "package_type", "priority", "weight_kg",
        "sender_name", "sender_phone", "sender_document",
        "recipient_name", "recipient_phone", "recipient_email",
        "address", "neighborhood", "city", "reference_point", "notes"
    ]

    sample_rows = [
        [
            "GTR-2026-901", "PAQUETE_MEDIANO", "ALTA", 2.5,
            "Almacén Éxito Cartagena", "3001112233", "890900608-9",
            "Dra. Mónica Patricia Orozco", "3158904561", "monica.orozco@gmail.com",
            "Cra. 3 # 8-15, Edificio Flamingo Apto 602", "Bocagrande", "Cartagena",
            "Frente a Droguería La Rebaja de la Cra 3", "Llamar al citófono 602 antes de subir"
        ],
        [
            "GTR-2026-902", "DOCUMENTO", "MEDIA", 0.4,
            "Notaría Tercera de Cartagena", "3104567890", "900123456-1",
            "Carlos Eduardo Valderrama", "3007654321", "carlos.valderrama@outlook.com",
            "Calle de la Media Luna # 10-45", "Getsemaní", "Cartagena",
            "A 50 metros del Parque Centenario", "Entregar en recepción con firma de recibido"
        ],
        [
            "GTR-2026-903", "REFRIGERADO", "ALTA", 4.8,
            "Mariscos del Caribe S.A.S.", "3189998877", "901456789-3",
            "Restaurante El Burlador", "3201239876", "pedidos@elburlador.com",
            "Calle Santo Domingo # 33-88", "Centro Histórico", "Cartagena",
            "Cerca a la Plaza Santo Domingo", "Mantener en cadena de frío - Urgente"
        ]
    ]

    if file_format == 'csv':
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(sample_headers)
        for r in sample_rows:
            writer.writerow(r)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="Plantilla_Importacion_GTR_Logistics.csv"'
        return response

    else:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Envíos GTR"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid") # Sky 600
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        ws.append(sample_headers)
        for col_num in range(1, len(sample_headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in sample_rows:
            ws.append(row)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Plantilla_Importacion_GTR_Logistics.xlsx"'
        return response


@csrf_exempt
def api_scan_barcode_qr(request):
    """
    Decodifica la información del escaneo de código de barras o QR de una etiqueta logística.
    Soporta formato JSON estructurado, delimitado por pipes/comas, o búsqueda directa por ID/Guía.
    """
    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No autenticado", "success": False}, status=401)

    if request.method != 'POST':
        return HttpResponseBadRequest()

    try:
        body = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest(json.dumps({"detail": "JSON inválido"}), content_type="application/json")

    raw_code = str(body.get('code') or body.get('raw_data') or '').strip()
    save_now = body.get('save_now', False)

    if not raw_code:
        return JsonResponse({"detail": "Código no proporcionado", "success": False}, status=400)

    # 1. Intentar buscar si ya existe en la base de datos por tracking code o ID
    if raw_code.isdigit():
        existing = Delivery.objects.filter(Q(tracking_code__iexact=raw_code) | Q(id=int(raw_code))).first()
    else:
        existing = Delivery.objects.filter(tracking_code__iexact=raw_code).first()
    if existing:
        return JsonResponse({
            "success": True,
            "source": "DATABASE_MATCH",
            "exists": True,
            "data": {
                "id": existing.id,
                "tracking_code": existing.tracking_code,
                "package_type": existing.package_type,
                "weight_kg": existing.weight_kg,
                "priority": existing.priority,
                "sender_name": existing.sender_name or "GTR Hub Central",
                "sender_phone": existing.sender_phone or "6056641234",
                "sender_document": existing.sender_document or "901234567-8",
                "recipient_name": existing.recipient_name,
                "recipient_phone": existing.recipient_phone,
                "recipient_email": existing.recipient_email or "",
                "address": existing.address,
                "neighborhood": existing.neighborhood or "Cartagena",
                "city": existing.city or "Cartagena",
                "reference_point": existing.reference_point or "",
                "status": existing.status,
                "notes": existing.notes or ""
            }
        })

    # 2. Intentar parsear como JSON incrustado en el QR
    decoded_dict = {}
    if raw_code.startswith('{') and raw_code.endswith('}'):
        try:
            parsed_json = json.loads(raw_code)
            decoded_dict = {
                "tracking_code": parsed_json.get('tracking') or parsed_json.get('tracking_code') or parsed_json.get('guia') or raw_code[:20],
                "package_type": (parsed_json.get('type') or parsed_json.get('package_type') or 'PAQUETE_MEDIANO').upper().replace(' ', '_'),
                "weight_kg": float(parsed_json.get('weight') or parsed_json.get('weight_kg') or 1.0),
                "priority": (parsed_json.get('priority') or 'MEDIA').upper(),
                "sender_name": parsed_json.get('sender') or parsed_json.get('sender_name') or 'Logística Integral Cartagena',
                "sender_phone": parsed_json.get('sender_phone') or '3001234567',
                "sender_document": parsed_json.get('sender_document') or '900987654-1',
                "recipient_name": parsed_json.get('recipient') or parsed_json.get('recipient_name') or 'Cliente Destinatario',
                "recipient_phone": parsed_json.get('phone') or parsed_json.get('recipient_phone') or '3000000000',
                "recipient_email": parsed_json.get('email') or parsed_json.get('recipient_email') or '',
                "address": parsed_json.get('address') or parsed_json.get('direccion') or 'Cartagena',
                "neighborhood": parsed_json.get('neighborhood') or parsed_json.get('barrio') or 'Centro Histórico',
                "city": parsed_json.get('city') or 'Cartagena',
                "reference_point": parsed_json.get('reference_point') or parsed_json.get('referencia') or '',
                "notes": parsed_json.get('notes') or ''
            }
        except Exception:
            pass

    # 3. Intentar parsear como formato delimitado por pipes: GUIA|TIPO|REMITENTE|DESTINATARIO|TEL|DIR|BARRIO|PESO|PRIORIDAD
    if not decoded_dict and '|' in raw_code:
        parts = [p.strip() for p in raw_code.split('|')]
        decoded_dict = {
            "tracking_code": parts[0] if len(parts) > 0 else f"GTR-{datetime.now().strftime('%Y%m%d%H%M')}",
            "package_type": parts[1].upper().replace(' ', '_') if len(parts) > 1 else 'PAQUETE_MEDIANO',
            "sender_name": parts[2] if len(parts) > 2 else 'GTR Logística S.A.S',
            "sender_phone": '3001234567',
            "sender_document": '900123456-1',
            "recipient_name": parts[3] if len(parts) > 3 else 'Cliente Destinatario',
            "recipient_phone": parts[4] if len(parts) > 4 else '3000000000',
            "recipient_email": '',
            "address": parts[5] if len(parts) > 5 else 'Cartagena',
            "neighborhood": parts[6] if len(parts) > 6 else 'Centro Histórico',
            "city": 'Cartagena',
            "weight_kg": float(parts[7]) if len(parts) > 7 and parts[7].replace('.', '', 1).isdigit() else 1.5,
            "priority": parts[8].upper() if len(parts) > 8 and parts[8].upper() in ['ALTA', 'MEDIA', 'BAJA'] else 'MEDIA',
            "reference_point": parts[9] if len(parts) > 9 else '',
            "notes": ''
        }

    # 4. Formato simple por defecto
    if not decoded_dict:
        import random
        decoded_dict = {
            "tracking_code": raw_code if len(raw_code) > 4 else f"GTR-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}",
            "package_type": 'PAQUETE_MEDIANO',
            "weight_kg": 1.2,
            "priority": 'MEDIA',
            "sender_name": 'Despacho Central GTR',
            "sender_phone": '6056641234',
            "sender_document": '901234567-8',
            "recipient_name": f"Destinatario ({raw_code[-4:]})",
            "recipient_phone": '3005551234',
            "recipient_email": '',
            "address": 'Calle 30 # 18-24',
            "neighborhood": 'Manga',
            "city": 'Cartagena',
            "reference_point": 'Frente al Edificio Puerto de Manga',
            "notes": f"Ingresado vía escáner de código: {raw_code}"
        }

    # Si se solicitó guardar directamente en base de datos
    if save_now:
        lat_f, lng_f = resolve_cartagena_coords(decoded_dict.get('neighborhood'), decoded_dict.get('address'))
        client_obj, _ = Client.objects.get_or_create(
            name=decoded_dict['recipient_name'],
            phone=decoded_dict['recipient_phone'],
            defaults={
                "address": decoded_dict['address'],
                "neighborhood": decoded_dict['neighborhood'],
                "city": decoded_dict.get('city', 'Cartagena'),
                "created_by_user": request.user
            }
        )
        delivery_obj = Delivery.objects.create(
            tracking_code=decoded_dict['tracking_code'],
            client=client_obj,
            recipient_name=decoded_dict['recipient_name'],
            recipient_phone=decoded_dict['recipient_phone'],
            recipient_email=decoded_dict.get('recipient_email') or None,
            sender_name=decoded_dict.get('sender_name'),
            sender_phone=decoded_dict.get('sender_phone'),
            sender_document=decoded_dict.get('sender_document'),
            address=decoded_dict['address'],
            neighborhood=decoded_dict['neighborhood'],
            city=decoded_dict.get('city', 'Cartagena'),
            latitude=lat_f,
            longitude=lng_f,
            reference_point=decoded_dict.get('reference_point'),
            package_type=decoded_dict.get('package_type', 'PAQUETE_MEDIANO'),
            weight_kg=decoded_dict.get('weight_kg', 1.0),
            priority=decoded_dict.get('priority', 'MEDIA'),
            notes=decoded_dict.get('notes'),
            status='PENDIENTE'
        )
        decoded_dict['id'] = delivery_obj.id

    return JsonResponse({
        "success": True,
        "source": "SCANNED_LABEL",
        "exists": False,
        "data": decoded_dict
    })


# ==============================================================================
# --- MÓDULO: NOVEDADES Y MONITOREO DE REPARTIDORES EN TIEMPO REAL ---
# ==============================================================================

@csrf_exempt
def api_driver_novelties_list_create(request):
    """
    Listar o registrar novedades e incidencias operativas de repartidores en tiempo real.
    """
    from core.models import DriverNovelty

    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No autenticado", "success": False}, status=401)

    if request.method == 'GET':
        driver_id = request.GET.get('driver_id')
        severity = request.GET.get('severity')
        resolved = request.GET.get('resolved')

        qs = DriverNovelty.objects.select_related('driver', 'delivery', 'resolved_by_user').all()
        if driver_id:
            qs = qs.filter(driver_id=driver_id)
        if severity:
            qs = qs.filter(severity=severity)
        if resolved is not None:
            qs = qs.filter(is_resolved=(resolved.lower() in ['true', '1', 'yes']))

        data = []
        for n in qs[:50]:
            data.append({
                "id": n.id,
                "driver_id": n.driver_id,
                "driver_name": n.driver.name,
                "driver_plate": n.driver.vehicle_plate,
                "delivery_id": n.delivery_id,
                "delivery_tracking": n.delivery.tracking_code if n.delivery else None,
                "novelty_type": n.novelty_type,
                "novelty_type_display": n.get_novelty_type_display(),
                "severity": n.severity,
                "title": n.title,
                "description": n.description,
                "latitude": n.latitude,
                "longitude": n.longitude,
                "photo_url": n.photo_url,
                "is_resolved": n.is_resolved,
                "resolved_at": n.resolved_at.isoformat() if n.resolved_at else None,
                "resolved_by": n.resolved_by_user.full_name if n.resolved_by_user else None,
                "created_at": n.created_at.isoformat()
            })
        return JsonResponse(data, safe=False)

    elif request.method == 'POST':
        try:
            body = json.loads(request.body.decode('utf-8'))
        except Exception:
            return HttpResponseBadRequest()

        driver_id = body.get('driver_id')
        if not driver_id and request.user.driver:
            driver_id = request.user.driver.id

        driver = Driver.objects.filter(id=driver_id).first()
        if not driver:
            return JsonResponse({"detail": "Repartidor no encontrado", "success": False}, status=404)

        title = body.get('title', '').strip() or 'Novedad Operativa en Ruta'
        description = body.get('description', '').strip()
        novelty_type = body.get('novelty_type', 'OTRO')
        severity = body.get('severity', 'MEDIA')
        delivery_id = body.get('delivery_id')

        lat_val = body.get('latitude', driver.current_lat)
        lng_val = body.get('longitude', driver.current_lng)

        photo_base64 = body.get('photo')
        photo_url = None
        if photo_base64:
            photo_url = save_base64_image(photo_base64, subfolder="novelties", prefix=f"nov_{driver.id}_")

        novelty = DriverNovelty.objects.create(
            driver=driver,
            delivery_id=delivery_id if delivery_id else None,
            novelty_type=novelty_type,
            severity=severity,
            title=title,
            description=description,
            latitude=lat_val,
            longitude=lng_val,
            photo_url=photo_url
        )

        log_audit_event(
            user=request.user,
            action="CREATE_DRIVER_NOVELTY",
            module="DISPATCH",
            ip_address=get_client_ip(request),
            details={"novelty_id": novelty.id, "driver": driver.name, "severity": severity}
        )

        return JsonResponse({
            "success": True,
            "id": novelty.id,
            "title": novelty.title,
            "created_at": novelty.created_at.isoformat()
        }, status=201)

    return HttpResponseBadRequest()


@csrf_exempt
def api_resolve_driver_novelty(request, novelty_id):
    """
    Marca una novedad operativa como resuelta.
    """
    from core.models import DriverNovelty

    if not request.user.is_authenticated:
        return JsonResponse({"detail": "No autenticado", "success": False}, status=401)

    if request.method != 'POST':
        return HttpResponseBadRequest()

    novelty = DriverNovelty.objects.filter(id=novelty_id).first()
    if not novelty:
        return JsonResponse({"detail": "Novedad no encontrada", "success": False}, status=404)

    novelty.is_resolved = True
    novelty.resolved_at = timezone.now()
    novelty.resolved_by_user = request.user
    novelty.save(update_fields=['is_resolved', 'resolved_at', 'resolved_by_user'])

    log_audit_event(
        user=request.user,
        action="RESOLVE_DRIVER_NOVELTY",
        module="DISPATCH",
        ip_address=get_client_ip(request),
        details={"novelty_id": novelty.id, "driver": novelty.driver.name}
    )

    return JsonResponse({"success": True, "message": "Novedad marcada como resuelta"})


@csrf_exempt
def api_driver_live_status(request):
    """
    Retorna el estado en vivo de toda la flota de repartidores para monitoreo en tiempo real.
    """
    from core.models import DriverNovelty

    drivers = Driver.objects.all()
    fleet = []

    for d in drivers:
        route = Route.objects.filter(driver=d).order_by('-created_at').first()
        stops = list(RouteStop.objects.filter(route=route).select_related('delivery').order_by('sequence_order')) if route else []
        delivered = sum(1 for s in stops if s.status == 'COMPLETADO' or (s.delivery and s.delivery.status == 'ENTREGADO'))
        in_transit = sum(1 for s in stops if s.status == 'EN_CAMINO' or (s.delivery and s.delivery.status == 'EN_CAMINO'))
        incidents = sum(1 for s in stops if s.status == 'INCIDENCIA' or (s.delivery and s.delivery.status == 'INCIDENCIA'))
        pending = max(0, len(stops) - delivered - in_transit - incidents)

        open_novelties = DriverNovelty.objects.filter(driver=d, is_resolved=False).count()

        fleet.append({
            "id": d.id,
            "name": d.name,
            "phone": d.phone,
            "vehicle_type": d.vehicle_type,
            "vehicle_plate": d.vehicle_plate,
            "capacity_kg": d.capacity_kg,
            "current_lat": d.current_lat or 10.4075,
            "current_lng": d.current_lng or -75.5342,
            "status": d.status,
            "route_code": route.code if route else None,
            "total_stops": len(stops),
            "delivered": delivered,
            "in_transit": in_transit,
            "incidents": incidents,
            "pending": pending,
            "progress_pct": round((delivered / len(stops) * 100)) if stops else 0,
            "open_novelties": open_novelties,
            "updated_at": d.updated_at.isoformat() if d.updated_at else timezone.now().isoformat()
        })

    return JsonResponse({"success": True, "fleet": fleet})

